

import logging
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
POS_FILL     = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL     = PatternFill("solid", fgColor="FFC7CE")
NEU_FILL     = PatternFill("solid", fgColor="FFEB9C")
WHITE_FONT   = Font(color="FFFFFF", bold=True)
BOLD_FONT    = Font(bold=True)
THIN         = Side(style="thin")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_header(ws, values: list, row: int, fill=HEADER_FILL):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)


def _sentiment_fill(s: str) -> PatternFill:
    return {"Positive": POS_FILL, "Negative": NEG_FILL}.get(s, NEU_FILL)


# ── Sheet 1: Channel Summary ──────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, channel_reports: dict):
    ws = wb.active
    ws.title = "Channel Summary"
    headers = ["Channel", "Top Topic", "Overall Mood", "Mood Score",
               "Content Style", "Key Themes", "Summary"]
    _write_header(ws, headers, row=1)

    for i, (channel, report) in enumerate(channel_reports.items(), 2):
        values = [
            channel,
            report.get("top_topic", ""),
            report.get("overall_mood", ""),
            report.get("mood_score", ""),
            report.get("content_style", ""),
            report.get("key_themes", ""),
            report.get("summary", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = ALT_FILL if i % 2 == 0 else PatternFill()
        # Colour the mood cell
        ws.cell(row=i, column=3).fill = _sentiment_fill(report.get("overall_mood", ""))

    _auto_width(ws)


# ── Sheet 2: Post Details ─────────────────────────────────────────────────────

def _sheet_posts(wb: Workbook, raw_posts: dict, post_analyses: dict):
    ws = wb.create_sheet("Post Details")
    headers = ["Channel", "Post #", "Post Preview (100 chars)",
               "Topic", "Sentiment", "Emotion", "Breaking?", "Keywords"]
    _write_header(ws, headers, row=1)

    row_idx = 2
    for channel, posts in raw_posts.items():
        analyses = post_analyses.get(channel, [])
        for i, (post, analysis) in enumerate(zip(posts, analyses), 1):
            values = [
                channel, i,
                post[:100] + ("…" if len(post) > 100 else ""),
                analysis.get("topic", ""),
                analysis.get("sentiment", ""),
                analysis.get("emotion", ""),
                analysis.get("is_breaking", ""),
                analysis.get("keywords", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
            # Colour sentiment cell
            ws.cell(row=row_idx, column=5).fill = _sentiment_fill(analysis.get("sentiment", ""))
            row_idx += 1

    _auto_width(ws)


# ── Sheet 3: Statistics ───────────────────────────────────────────────────────

def _sheet_stats(wb: Workbook, post_analyses: dict):
    ws = wb.create_sheet("Statistics")

    all_topics, all_sentiments, all_emotions = [], [], []
    for analyses in post_analyses.values():
        for a in analyses:
            all_topics.append(a.get("topic", "Unknown"))
            all_sentiments.append(a.get("sentiment", "Unknown"))
            all_emotions.append(a.get("emotion", "Unknown"))

    total = max(len(all_topics), 1)

    def write_table(title: str, counter: Counter, start_row: int) -> int:
        ws.cell(row=start_row, column=1, value=title).font = WHITE_FONT
        ws.cell(row=start_row, column=1).fill = SUBHEAD_FILL
        ws.cell(row=start_row, column=2, value="Count").font = WHITE_FONT
        ws.cell(row=start_row, column=2).fill = SUBHEAD_FILL
        ws.cell(row=start_row, column=3, value="%").font = WHITE_FONT
        ws.cell(row=start_row, column=3).fill = SUBHEAD_FILL
        r = start_row + 1
        for k, v in counter.most_common():
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v)
            ws.cell(row=r, column=3, value=f"{v * 100 // total}%")
            r += 1
        return r + 1

    row = 1
    row = write_table("📊 Topics", Counter(all_topics), row)
    row = write_table("😊 Sentiments", Counter(all_sentiments), row)
    row = write_table("💡 Emotions", Counter(all_emotions), row)

    _auto_width(ws)


# ── Main ──────────────────────────────────────────────────────────────────────

def create_excel_report(result_path: str, raw_posts: dict,
                        post_analyses: dict, channel_reports: dict):
    wb = Workbook()
    _sheet_summary(wb, channel_reports)
    _sheet_posts(wb, raw_posts, post_analyses)
    _sheet_stats(wb, post_analyses)
    wb.save(result_path)
    logging.info(f"[export] Excel report saved: {result_path}")
